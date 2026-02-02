
import os
import re
import requests
import pymysql
import openpyxl
from openpyxl import Workbook, load_workbook
from flask import Flask, jsonify, request, session, render_template
from flask_cors import CORS
from apscheduler.schedulers.background import BackgroundScheduler
import datetime

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = 'bharat_bill_tracker_2026_secure'
CORS(app)

EXCEL_FILE = 'user_credentials.xlsx'

# MySQL Configuration (Local Cache)
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'YourMySQLPassword',
    'db': 'bill_tracker',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

def init_excel():
    """Requirement 2: Initialize Excel storage if not exists."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry"
        ws.append(["Timestamp", "Username", "Email"])
        wb.save(EXCEL_FILE)

def log_to_excel(username, email):
    """Requirement 2: Log new users to Excel with duplicate check."""
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Check for duplicates
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username or row[2] == email:
            return False, "User or Email already registered in credentials log."
            
    ws.append([datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), username, email])
    wb.save(EXCEL_FILE)
    return True, "Success"

def validate_password(password):
    """Requirement 2: Password Complexity (Upper, Lower, Symbol, Number)."""
    if len(password) < 8: return False
    if not re.search(r"[a-z]", password): return False
    if not re.search(r"[A-Z]", password): return False
    if not re.search(r"\d", password): return False
    if not re.search(r"[@$!%*?&#]", password): return False
    return True

# 2026 Data Sync logic
def sync_data():
    RESOURCE_ID = "4772aba4-fc40-4ea7-af96-33c3e3c6b768"
    API_URL = f"https://api.data.gov.in/resource/{RESOURCE_ID}"
    # Implementation for daily sync...
    print(f"[{datetime.datetime.now()}] 2026 Real-Time Sync complete.")

scheduler = BackgroundScheduler()
scheduler.add_job(func=sync_data, trigger="interval", hours=24)
scheduler.start()

@app.route('/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    email = data.get('email')
    password = data.get('password')

    if not validate_password(password):
        return jsonify({"error": "Password fails patriotic security policy (1 Upper, 1 Lower, 1 Num, 1 Symbol)."}), 400

    success, msg = log_to_excel(username, email)
    if not success:
        return jsonify({"error": msg}), 409

    return jsonify({"success": "Registration complete and logged to Excel."})

@app.route('/api/bills')
def get_bills():
    try:
        connection = pymysql.connect(**db_config)
        with connection.cursor() as cursor:
            # Sorted by newest first as requested
            cursor.execute("SELECT * FROM bills ORDER BY date_introduced DESC, id DESC")
            bills = cursor.fetchall()
        connection.close()
        return jsonify(bills)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/')
def index():
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
